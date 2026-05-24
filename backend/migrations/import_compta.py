"""
Import financial data from Compta.xlsx into encrypted SQLite database.

This script reads the Excel workbook and populates the database with:
- Cash accounts (currencies)
- Investments (stocks, ETFs)
- Properties (real estate with amortization)
- Tax data
- Monthly snapshots (historical AUM)
"""

import sys
from pathlib import Path
from datetime import datetime, date
from openpyxl import load_workbook
import logging

# Add backend to path
backend_dir = Path(__file__).parent.parent.parent
sys.path.insert(0, str(backend_dir))

from backend.database import SessionLocal, create_tables
from backend.db import (
    CashAccount, Investment, Property, PropertyAmortization,
    PropertyCashflow, PropertyEquity, MonthlySnapshot, TaxData, Forecast
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ComptaImporter:
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.wb = load_workbook(self.excel_path)
        self.db = SessionLocal()

    def get_sheet(self, name: str):
        """Get worksheet by name"""
        if name not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{name}' not found in workbook")
        return self.wb[name]

    def cell_value(self, sheet, row, col):
        """Safely get cell value"""
        try:
            cell = sheet.cell(row, col)
            return cell.value
        except:
            return None

    def import_all(self):
        """Run all import steps"""
        try:
            logger.info("Creating database tables...")
            create_tables()

            logger.info("Importing cash accounts...")
            self.import_cash_accounts()

            logger.info("Importing investments...")
            self.import_investments()

            logger.info("Importing properties...")
            self.import_properties()

            logger.info("Importing tax data...")
            self.import_tax_data()

            logger.info("Importing historical snapshots...")
            self.import_snapshots()

            logger.info("✓ Import completed successfully")
            self.db.commit()

        except Exception as e:
            logger.error(f"Import failed: {e}")
            self.db.rollback()
            raise
        finally:
            self.db.close()

    def import_cash_accounts(self):
        """Import cash accounts from Overview sheet"""
        ws = self.get_sheet("Overview")

        # Based on Compta.xlsx structure:
        # Row 2 has: SGD rate, GBP rate, etc.
        # Row 3 onwards has: HKD, USD, JPY, AUD rates

        currencies_data = [
            ("EUR", 1, 1),      # EUR always 1.0
            ("SGD", self.cell_value(ws, 2, 1), 2),
            ("GBP", self.cell_value(ws, 2, 2), 2),
            ("HKD", self.cell_value(ws, 3, 1), 3),
            ("USD", self.cell_value(ws, 4, 1), 4),
            ("JPY", self.cell_value(ws, 5, 1), 5),
            ("AUD", self.cell_value(ws, 6, 1), 6),
        ]

        for currency, rate, row in currencies_data:
            if rate is None or rate == 0:
                continue

            # Get account details from Overview sheet
            # For now, create basic cash accounts (values can be updated from actual holdings)
            account = CashAccount(
                currency=currency,
                amount=0,  # Will update from actual holdings
                eur_amount=0,
                exchange_rate=rate,
                account_name=f"Main {currency}",
                account_type="cash",
            )
            self.db.add(account)

        self.db.commit()

    def import_investments(self):
        """Import investments from Overview sheet"""
        ws = self.get_sheet("Overview")

        # Investment section starts around column 10
        # Symbols: DBS, HSBC HK, HSBC SG, Interactive Broker
        investments_info = [
            ("DBS", "stock", 10, 3),  # Row 3
            ("HSBC HK", "stock", 10, 4),  # Row 4
            ("HSBC SG", "stock", 10, 5),  # Row 5
        ]

        for symbol, inv_type, col, row in investments_info:
            name = self.cell_value(ws, row, col)
            currency = self.cell_value(ws, row, col + 1)
            quantity = self.cell_value(ws, row, col + 2)
            cost_basis = self.cell_value(ws, row, col + 3)
            eur_amount = self.cell_value(ws, row, col + 4)
            yield_val = self.cell_value(ws, row, col + 5)

            if symbol and quantity:
                inv = Investment(
                    symbol=symbol,
                    name=name,
                    investment_type=inv_type,
                    quantity=float(quantity) if quantity else 0,
                    cost_basis=float(cost_basis) if cost_basis else 0,
                    currency=currency if currency else "SGD",
                    eur_amount=float(eur_amount) if eur_amount else 0,
                    current_price=0,
                    yield_pct=float(yield_val) if yield_val else None,
                )
                self.db.add(inv)

        self.db.commit()

    def import_properties(self):
        """Import properties and amortization from property sheets"""
        property_sheets = ["Roses", "Wagner", "Parmentier", "Boulets", "Larmor Plage"]

        for sheet_name in property_sheets:
            if sheet_name not in self.wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found, skipping")
                continue

            ws = self.get_sheet(sheet_name)

            # Create property record
            # Get purchase price and current value from sheet
            purchase_price = self.cell_value(ws, 2, 17) or 0  # Cash initial

            prop = Property(
                name=sheet_name,
                property_type="rental" if sheet_name != "Larmor Plage" else "investment",
                purchase_price=float(purchase_price) if purchase_price else 0,
                current_value=float(purchase_price) if purchase_price else 0,
                currency="EUR",
            )
            self.db.add(prop)
            self.db.flush()  # Get the property ID

            # Import amortization data (monthly schedule)
            for row in range(4, ws.max_row + 1):
                month = self.cell_value(ws, row, 1)
                date_val = self.cell_value(ws, row, 2)
                principal_paid = self.cell_value(ws, row, 3)
                interest_paid = self.cell_value(ws, row, 4)
                capital_left = self.cell_value(ws, row, 5)

                if month is None or capital_left is None:
                    break  # End of data

                # Convert Excel serial date to Python date
                if isinstance(date_val, (int, float)):
                    # Excel epoch: Jan 1, 1900 (but with 1-day bug)
                    try:
                        excel_date = datetime(1900, 1, 1) + \
                                    __import__('datetime').timedelta(days=date_val - 2)
                        date_obj = excel_date.date()
                    except:
                        date_obj = date.today()
                else:
                    date_obj = date.today()

                amort = PropertyAmortization(
                    property_id=prop.id,
                    month=int(month) if month else 0,
                    date=date_obj,
                    capital_left=float(capital_left) if capital_left else 0,
                    interest_paid=float(interest_paid) if interest_paid else 0,
                    principal_paid=float(principal_paid) if principal_paid else 0,
                    total_paid=(float(principal_paid) if principal_paid else 0) +
                              (float(interest_paid) if interest_paid else 0),
                )
                self.db.add(amort)

        self.db.commit()

    def import_tax_data(self):
        """Import tax information from Tax sheet"""
        if "Tax" not in self.wb.sheetnames:
            logger.warning("Tax sheet not found, skipping")
            return

        ws = self.get_sheet("Tax")

        # Tax data structure:
        # Row 2: headers (years)
        # Row 3: Income
        # Row 4-6: Reliefs
        # Row 7: Taxable income
        # Row 8: Pre-rebate tax
        # Row 10: Final tax

        years_row = ws[2]
        years = []
        for col in range(5, 12):  # Columns E-K typically
            year_val = self.cell_value(ws, 2, col)
            if year_val and isinstance(year_val, int) and 2000 <= year_val <= 2100:
                years.append((col, year_val))

        for col, year in years:
            income = self.cell_value(ws, 3, col) or 0
            taxable = self.cell_value(ws, 7, col) or 0
            pre_rebate = self.cell_value(ws, 8, col) or 0
            total_tax = self.cell_value(ws, 10, col) or 0

            tax_data = TaxData(
                year=int(year),
                income=float(income) if income else 0,
                taxable_income=float(taxable) if taxable else 0,
                pre_rebate_tax=float(pre_rebate) if pre_rebate else 0,
                total_tax=float(total_tax) if total_tax else 0,
            )
            self.db.add(tax_data)

        self.db.commit()

    def import_snapshots(self):
        """Import historical AUM snapshots from 'Historique projection' sheet"""
        if "Historique projection" not in self.wb.sheetnames:
            logger.warning("Historique projection sheet not found, skipping")
            return

        ws = self.get_sheet("Historique projection")

        # Structure:
        # Row 1: Headers (years: 2020, 2021, etc.)
        # Row 2+: Date and AUM values

        for row in range(2, min(ws.max_row + 1, 50)):  # First 50 rows
            date_val = self.cell_value(ws, row, 2)
            aum = self.cell_value(ws, row, 3)

            if date_val is None or aum is None:
                continue

            # Convert Excel serial date
            if isinstance(date_val, (int, float)):
                try:
                    excel_date = datetime(1900, 1, 1) + \
                                __import__('datetime').timedelta(days=date_val - 2)
                    date_obj = excel_date.date()
                except:
                    continue
            else:
                continue

            snapshot = MonthlySnapshot(
                date=date_obj,
                total_assets_eur=float(aum) if aum else 0,
                total_debt_eur=0,  # Would need to extract from property sheets
                net_equity_eur=float(aum) if aum else 0,
            )
            self.db.add(snapshot)

        self.db.commit()

    def close(self):
        """Close database connection"""
        self.db.close()

def main():
    """Main entry point"""
    excel_path = Path(__file__).parent.parent.parent / "Data" / "Compta.xlsx"

    if not excel_path.exists():
        logger.error(f"Excel file not found: {excel_path}")
        return 1

    logger.info(f"Starting import from {excel_path}")

    importer = ComptaImporter(str(excel_path))
    try:
        importer.import_all()
        logger.info("✓ Data import successful!")
        return 0
    except Exception as e:
        logger.error(f"✗ Import failed: {e}")
        import traceback
        traceback.print_exc()
        return 1
    finally:
        importer.close()

if __name__ == "__main__":
    sys.exit(main())
